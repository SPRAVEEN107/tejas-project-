# group_photos_attendance.py
# Scans a folder of group photos, recognizes faces using ArcFace, and logs attendance to an Excel file.
# Usage:
#   python group_photos_attendance.py
#   python group_photos_attendance.py group_photos

import os
import sys
import glob
import warnings
from datetime import datetime

import cv2
import numpy as np
import insightface
from pymongo import MongoClient
from numpy.linalg import norm
import openpyxl

# ---------------------------
# CONFIG
# ---------------------------
# Default folder to read images from
DEFAULT_FOLDER = "group_photos"

# MongoDB Atlas URI (update if needed)
MONGO_URI = "mongodb+srv://praveen:yourstrongpassword123@attendancedb.tpwmjgs.mongodb.net/?retryWrites=true&w=majority"
DB_NAME = "AttendanceDB"
STUDENTS_COLL = "students"

# ArcFace threshold and device
ARC_THRESHOLD = 0.38   # adjust later if needed
CTX_ID = -1            # -1 = CPU, 0 = GPU

# allowed image extensions
IMG_EXTS = ("*.jpg", "*.jpeg", "*.png", "*.bmp")

# ---------------------------
# SILENCE NOISY WARNINGS (optional)
# ---------------------------
warnings.filterwarnings("ignore")
os.environ["ORT_LOGGING_LEVEL"] = "ERROR"
os.environ["INSIGHTFACE_DEBUG"] = "0"

# ---------------------------
# HELPER: cosine similarity
# ---------------------------
def cosine_similarity(a, b):
    a = np.array(a, dtype=float)
    b = np.array(b, dtype=float)
    return float(np.dot(a, b) / (norm(a) * norm(b)))

# ---------------------------
# LOAD ARCFACE MODEL
# ---------------------------
print("Loading ArcFace model (buffalo_l)...")
model = insightface.app.FaceAnalysis(name="buffalo_l")
model.prepare(ctx_id=CTX_ID, det_size=(640, 640))
print("ArcFace model ready.")

# ---------------------------
# LOAD KNOWN ENCODINGS FROM MONGODB
# ---------------------------
def load_known_encodings():
    client = MongoClient(MONGO_URI)
    db = client[DB_NAME]
    col = db[STUDENTS_COLL]

    names = []
    rolls = []
    encodings = []

    for doc in col.find({"model": "arcface"}):
        names.append(doc.get("name", "Unknown"))
        rolls.append(str(doc.get("rollNo", "")))
        enc = doc.get("face_encoding", None)
        if enc and len(enc) >= 16:  # basic sanity
            encodings.append(np.array(enc, dtype=float))
        else:
            encodings.append(None)
    # filter out any None encodings together
    filtered = [(n, r, e) for n, r, e in zip(names, rolls, encodings) if e is not None]
    if not filtered:
        return [], [], []
    names_f, rolls_f, encs_f = zip(*filtered)
    return list(names_f), list(rolls_f), list(encs_f)

names_list, rolls_list, known_embeddings = load_known_encodings()
print(f"Loaded {len(known_embeddings)} ArcFace encodings from DB.")

# ---------------------------
# PREPARE TODAY'S ATTENDANCE FILE
# ---------------------------
def attendance_filename_for_date(date_str):
    return f"attendance_{date_str}.xlsx"

today_str = datetime.now().strftime("%Y-%m-%d")
attendance_file = attendance_filename_for_date(today_str)

if not os.path.exists(attendance_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Roll No", "Name", "Timestamp"])
    wb.save(attendance_file)
    print(f"Created attendance file: {attendance_file}")
else:
    print(f"Using existing attendance file: {attendance_file}")

# load workbook and current marks
wb = openpyxl.load_workbook(attendance_file)
ws = wb.active
marked_today = set(str(row[0]) for row in ws.iter_rows(min_row=2, values_only=True) if row[0] is not None)

# ---------------------------
# PROCESS FOLDER
# ---------------------------
def process_folder(folder):
    if not os.path.exists(folder):
        print(f"❌ Folder not found: {folder}")
        return

    # collect files
    files = []
    for ext in IMG_EXTS:
        files.extend(glob.glob(os.path.join(folder, ext)))
    files = sorted(files)
    if not files:
        print(f"❌ No images found inside folder: {folder}")
        return

    print(f"Found {len(files)} image(s) in '{folder}'. Processing...")

    for filepath in files:
        basename = os.path.basename(filepath)
        print(f"\n--- Processing: {basename}")

        # read image
        img = cv2.imread(filepath)
        if img is None:
            print(f"  ❌ Could not open/read file: {filepath} — skipping.")
            continue

        # detect faces
        faces = model.get(img)
        print(f"  Detected {len(faces)} face(s).")

        # keep track if anything changed in this image
        any_marked = False

        for face in faces:
            bbox = face.bbox.astype(int)
            x1, y1, x2, y2 = bbox
            emb = face.embedding
            # normalize
            emb = emb / np.linalg.norm(emb)

            # compute similarities
            if not known_embeddings:
                print("  ❌ No known embeddings loaded — cannot match.")
                label = "Unknown"
                cv2.rectangle(img, (x1, y1), (x2, y2), (0,0,255), 2)
                cv2.putText(img, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0,0,255), 2)
                continue

            sims = [cosine_similarity(emb, k) for k in known_embeddings]
            best_idx = int(np.argmax(sims))
            best_score = sims[best_idx]

            if best_score >= ARC_THRESHOLD:
                name = names_list[best_idx]
                roll = rolls_list[best_idx]
                label = f"{name} ({roll}) {best_score:.2f}"
                cv2.rectangle(img, (x1, y1), (x2, y2), (0,255,0), 2)
                cv2.putText(img, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0,255,0), 2)
                print(f"  ✓ Recognized: {name} ({roll}) score={best_score:.3f}")

                # mark attendance if not already
                if roll not in marked_today:
                    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    ws.append([roll, name, timestamp])
                    marked_today.add(roll)
                    any_marked = True
                    print(f"    → Marked present at {timestamp}")
                else:
                    print("    → Already marked today (skipping)")

            else:
                label = "Unknown"
                cv2.rectangle(img, (x1, y1), (x2, y2), (0,0,255), 2)
                cv2.putText(img, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0,0,255), 2)
                print(f"  ✗ No confident match (best score={best_score:.3f})")

        # save labeled output image
        out_name = os.path.join(folder, f"out_{basename}")
        success = cv2.imwrite(out_name, img)
        if success:
            print(f"  ➜ Saved labeled image: {out_name}")
        else:
            print("  ❌ Failed to save labeled image.")

        # save workbook after each processed image where new marks were added
        if any_marked:
            wb.save(attendance_file)
            print(f"  ➜ Attendance file updated: {attendance_file}")

    # final save
    wb.save(attendance_file)
    print(f"\nProcessing complete. Final attendance saved to: {attendance_file}")


# ---------------------------
# MAIN
# ---------------------------
if __name__ == "__main__":
    folder = DEFAULT_FOLDER
    if len(sys.argv) > 1:
        folder = sys.argv[1].strip()

    print(f"Folder set to: {folder}")
    process_folder(folder)

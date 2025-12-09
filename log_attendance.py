import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from pymongo import MongoClient

# --------------------------------------------------
# CONFIGURATION
# --------------------------------------------------

MONGO_URI = "mongodb+srv://praveen:yourstrongpassword123@attendancedb.tpwmjgs.mongodb.net/?retryWrites=true&w=majority"
DB_NAME = "AttendanceDB"
ATTENDANCE_FOLDER = "."   # Current folder


# --------------------------------------------------
# LOAD ALL STUDENTS FROM MONGODB
# --------------------------------------------------

def load_all_students():
    client = MongoClient(MONGO_URI)
    db = client[DB_NAME]
    students_col = db["students"]

    students = {}
    for doc in students_col.find():
        roll = str(doc.get("rollNo"))
        name = doc.get("name", "Unknown")
        students[roll] = name

    return students


# --------------------------------------------------
# GET TODAY'S ATTENDANCE FILE
# --------------------------------------------------

def get_attendance_file():
    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"attendance_{today}.xlsx"
    filepath = os.path.join(ATTENDANCE_FOLDER, filename)
    return filepath


# --------------------------------------------------
# READ PRESENT STUDENTS FOR SELECTED DATE
# --------------------------------------------------

def get_present_students(attendance_file):
    present = set()

    if not os.path.exists(attendance_file):
        return present  # File missing = no present students

    wb = load_workbook(attendance_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        roll, name, timestamp = row
        present.add(str(roll))

    return present


# --------------------------------------------------
# CREATE EXCEL REPORT
# --------------------------------------------------

def create_excel(filename, header, data_rows):
    wb = Workbook()
    ws = wb.active
    ws.append(header)

    for row in data_rows:
        ws.append(row)

    wb.save(filename)
    print(f"📁 Report saved: {filename}")


# --------------------------------------------------
# MAIN FUNCTION TO GENERATE PRESENT OR ABSENT REPORT
# --------------------------------------------------

def generate_final_sheet():
    print("\n-------------------------------------")
    print("Select Report Type:")
    print("1 → Present Students")
    print("2 → Absent Students")
    print("-------------------------------------")

    choice = input("Enter your choice (1 or 2): ").strip()

    # Load all students from DB
    students = load_all_students()

    # Get today's attendance file
    attendance_file = get_attendance_file()

    # Load present students
    present_students = get_present_students(attendance_file)

    if choice == "1":
        # PRESENT LIST
        data = []
        for roll in sorted(present_students):
            data.append([roll, students.get(roll, "Unknown")])

        create_excel("Present_List.xlsx", ["Roll No", "Name"], data)

    elif choice == "2":
        # ABSENT LIST
        all_rolls = set(students.keys())
        absent = sorted(all_rolls - present_students)

        data = []
        for roll in absent:
            data.append([roll, students.get(roll, "Unknown")])

        create_excel("Absent_List.xlsx", ["Roll No", "Name"], data)

    else:
        print("❌ Invalid choice. No report generated.")


# --------------------------------------------------
# RUN WHEN EXECUTED DIRECTLY
# --------------------------------------------------

if __name__ == "__main__":
    generate_final_sheet()
